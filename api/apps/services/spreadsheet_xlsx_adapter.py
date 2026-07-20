#
#  Copyright 2026 The InfiniFlow Authors. All Rights Reserved.
#
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
#
"""
xlsx ↔ Univer IWorkbookData adapter.

Storage format change (M1 of spreadsheet enhancement):
  collaboration_document.content now stores Univer-native IWorkbookData JSON
  instead of the legacy `{sheets:[{name,data,colWidths}]}` grid format.

This module owns ALL openpyxl ↔ Univer mapping logic so the rest of the
codebase only sees two stable entry points:
  - xlsx_to_workbook_data(bytes, doc_id) -> dict
  - workbook_data_to_xlsx(dict)         -> bytes

Scope (M1 = MVP):
  - cell values, formulas, booleans, numbers, datetimes (as ISO strings)
  - sheet names / order / visibility
  - column widths (openpyxl char-unit ↔ Univer pixel-unit)
  - row heights (openpyxl points ↔ Univer pixels)
  - hidden rows / columns
  - merged cells
  - freeze panes
  - sheet RTL

Deferred to later phases:
  P3: cell styles (font / fill / border / alignment / number format)
  P4: floating images (MinIO-backed)
  P5: comments / conditional formatting / data validation
  P6: chart placeholders
"""

import io
import logging
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# ── Hard limits ─────────────────────────────────────────────────────────
# Enforced up-front in xlsx_to_workbook_data to prevent pathological files
# from exhausting memory (e.g. a sheet with 1M empty rows that openpyxl
# happily expands into real XML). Anything larger should be split or
# pre-processed by the caller.
MAX_XLSX_SIZE = 10 * 1024 * 1024  # 10 MB
MAX_ROWS = 100_000
MAX_COLS = 1000
# Reserved for P4 — declared here so callers can reference it.
MAX_IMAGES = 50

# Conversion constants. openpyxl uses "character width" for columns and
# points for rows; Univer uses CSS pixels at 96 DPI. These multipliers are
# the canonical conversions used throughout the Office ecosystem.
_PX_PER_CHAR = 7.0     # approx width of one character in Calibri 11
_COL_PADDING = 5.0     # default cell left+right padding in pixels
_PX_PER_POINT = 1.33   # 96 / 72
_MIN_COL_W = 20.0
_MIN_ROW_H = 19.0

logger = logging.getLogger(__name__)


# ── Exceptions ──────────────────────────────────────────────────────────

class XlsxTooLargeError(ValueError):
    """Raised when xlsx exceeds MAX_XLSX_SIZE / MAX_ROWS / MAX_COLS."""

    def __init__(self, reason: str, code: str = "FILE_TOO_LARGE"):
        super().__init__(reason)
        self.code = code


# ── Public API ──────────────────────────────────────────────────────────

def xlsx_to_workbook_data(file_bytes: bytes, doc_id: str) -> dict:
    """Parse .xlsx bytes → Univer IWorkbookData.

    Raises XlsxTooLargeError on size/row/col/image violations.
    Images found in the workbook are uploaded to MinIO bucket "collaboration"
    under key `docs/{doc_id}/images/{uuid}.{ext}`. The drawing params reference
    these via the proxy URL `/api/v1/collaboration/documents/{doc_id}/assets/{uuid}`.
    The frontend rewrites the URL with `?token=<jwt>` at load time.
    """
    size = len(file_bytes)
    if size > MAX_XLSX_SIZE:
        raise XlsxTooLargeError(
            f"文件过大：{size // 1024 // 1024}MB 超过 {MAX_XLSX_SIZE // 1024 // 1024}MB 限制",
            code="FILE_TOO_LARGE",
        )

    # read_only=False is required for image/chart extraction (P4/P6).
    # keep_links=False avoids pulling external workbook references that
    # would otherwise produce #REF! errors during round-trip.
    wb = load_workbook(
        io.BytesIO(file_bytes),
        read_only=False,
        data_only=False,
        keep_links=False,
    )
    try:
        return _workbook_to_univer(wb, doc_id)
    finally:
        try:
            wb.close()
        except Exception:
            pass


def workbook_data_to_xlsx(content: dict) -> bytes:
    """Convert IWorkbookData (or legacy format) → .xlsx bytes.

    Accepts both:
      - New format: full IWorkbookData with sheetOrder + sheets dict
      - Legacy format: {sheets: [{name, data, colWidths}], activeSheet}
    """
    buf = io.BytesIO()
    wb = Workbook()

    sheets = content.get("sheets")
    if isinstance(sheets, list):
        _write_legacy_sheets(wb, sheets)
    else:
        _write_univer_sheets(
            wb,
            sheets or {},
            content.get("sheetOrder", []),
            styles=content.get("styles") or {},
        )

    wb.save(buf)
    return buf.getvalue()


def _apply_style_to_cell(cell, style: dict) -> None:
    """Apply a Univer IStyleData dict to an openpyxl cell (inverse of _build_style)."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    # Font
    font_kwargs: dict[str, Any] = {}
    if "ff" in style:
        font_kwargs["name"] = style["ff"]
    if "fs" in style:
        font_kwargs["size"] = style["fs"]
    if style.get("bl"):
        font_kwargs["bold"] = True
    if style.get("it"):
        font_kwargs["italic"] = True
    ul = style.get("ul")
    if isinstance(ul, dict) and ul.get("s"):
        font_kwargs["underline"] = "double" if ul.get("t") == "double" else "single"
    st = style.get("st")
    if isinstance(st, dict) and st.get("s"):
        font_kwargs["strike"] = True
    if isinstance(style.get("cl"), dict) and style["cl"].get("rgb"):
        # Univer stores "#RRGGBB"; openpyxl wants aRGB "FFRRGGBB"
        hex_part = str(style["cl"]["rgb"]).lstrip("#").upper()
        if len(hex_part) == 6:
            font_kwargs["color"] = "FF" + hex_part
    if font_kwargs:
        cell.font = Font(**font_kwargs)

    # Fill
    bg = style.get("bg")
    if isinstance(bg, dict) and bg.get("rgb"):
        hex_part = str(bg["rgb"]).lstrip("#").upper()
        if len(hex_part) == 6:
            cell.fill = PatternFill(fill_type="solid", fgColor="FF" + hex_part)

    # Border — Univer uses t/b/l/r; openpyxl uses top/bottom/left/right
    bd = style.get("bd")
    if isinstance(bd, dict) and bd:
        side_kwargs: dict[str, Any] = {}
        for u_key, opx_key in (("t", "top"), ("b", "bottom"), ("l", "left"), ("r", "right")):
            side_data = bd.get(u_key)
            if not isinstance(side_data, dict):
                continue
            s_style = side_data.get("s") or "thin"
            s_color = None
            c = side_data.get("c")
            if isinstance(c, dict) and c.get("rgb"):
                hex_part = str(c["rgb"]).lstrip("#").upper()
                if len(hex_part) == 6:
                    s_color = "FF" + hex_part
            side_kwargs[opx_key] = Side(style=s_style, color=s_color)
        if side_kwargs:
            cell.border = Border(**side_kwargs)

    # Alignment
    align_kwargs: dict[str, Any] = {}
    ht = style.get("ht")
    if ht in (1, 2, 3, 4, 7):
        align_kwargs["horizontal"] = {1: "left", 2: "center", 3: "right", 4: "justify", 7: "distributed"}.get(ht)
    vt = style.get("vt")
    if vt in (1, 2, 3, 4, 7):
        align_kwargs["vertical"] = {2: "top", 1: "center", 3: "bottom", 4: "justify", 7: "distributed"}.get(vt)
    if style.get("tb"):
        align_kwargs["wrap_text"] = True
    if "rt" in style:
        align_kwargs["text_rotation"] = int(style["rt"])
    if align_kwargs:
        cell.alignment = Alignment(**align_kwargs)

    # Number format
    n = style.get("n")
    if isinstance(n, str) and n:
        cell.number_format = n


# ── openpyxl → Univer ──────────────────────────────────────────────────

def _workbook_to_univer(wb, doc_id: str) -> dict:
    sheets_univer: dict[str, dict] = {}
    sheet_order: list[str] = []
    styles_registry = _StyleRegistry()
    # Per-sheet drawing maps: { sheetId: { drawingId: IDrawingParam } }
    all_drawings: dict[str, dict[str, dict]] = {}
    total_images = 0

    for idx, ws in enumerate(wb.worksheets):
        sheet_id = f"sheet_{idx}"
        sheet_order.append(sheet_id)
        sheets_univer[sheet_id] = _worksheet_to_univer(
            ws, sheet_id, idx, styles_registry
        )
        # Extract floating images (P4) — uploads to MinIO + builds drawing params.
        drawings, img_count = _read_drawings(ws, doc_id, sheet_id)
        if drawings:
            all_drawings[sheet_id] = drawings
            total_images += img_count
        # Extract charts (P6) — render as PNG placeholder + treat as image drawing.
        # Real chart rendering requires Pro license; without it we embed a labeled
        # PNG so users at least see WHERE a chart was and what type it was.
        chart_drawings, chart_count = _read_charts(ws, doc_id, sheet_id)
        if chart_drawings:
            if all_drawings.get(sheet_id):
                all_drawings[sheet_id].update(chart_drawings)
            else:
                all_drawings[sheet_id] = chart_drawings
            total_images += chart_count

    if total_images > MAX_IMAGES:
        raise XlsxTooLargeError(
            f"图片过多：{total_images} 张（上限 {MAX_IMAGES}）",
            code="TOO_MANY_IMAGES",
        )

    if not sheets_univer:
        sid = "sheet_0"
        sheet_order = [sid]
        sheets_univer[sid] = _blank_sheet(sid, "Sheet1")

    # Pack drawings into SHEET_DRAWING_PLUGIN resource.
    # Univer's runtime stores drawings as drawingManagerData[unitId][subUnitId].data[drawingId],
    # so the persisted JSON must match that shape: { sheetId: { "data": { drawId: param } } }.
    # Wrapping in ".data" is critical — Univer's removeDrawingDataForUnit and
    # initializeNotification access subUnit.data directly and crash with
    # Object.keys(null) if the wrapper is missing.
    resources: list[dict] = []
    if all_drawings:
        import json
        wrapped = {sid: {"data": d} for sid, d in all_drawings.items()}
        resources.append({
            "name": "SHEET_DRAWING_PLUGIN",
            "data": json.dumps(wrapped, ensure_ascii=False),
        })

    # Pack cell notes (P5) into SHEET_NOTE_PLUGIN resource.
    # Univer stores notes as { sheetId: { row: { col: ISheetNote } } }.
    all_notes: dict[str, dict[str, dict[str, dict]]] = {}
    for idx, ws in enumerate(wb.worksheets):
        sheet_id = f"sheet_{idx}"
        notes = _read_notes(ws)
        if notes:
            all_notes[sheet_id] = notes
    if all_notes:
        import json
        resources.append({
            "name": "SHEET_NOTE_PLUGIN",
            "data": json.dumps(all_notes, ensure_ascii=False),
        })

    # Pack data validations (P5) into SHEET_DATA_VALIDATION_PLUGIN resource.
    all_dvs: dict[str, list[dict]] = {}
    for idx, ws in enumerate(wb.worksheets):
        sheet_id = f"sheet_{idx}"
        dvs = _read_data_validations(ws)
        if dvs:
            all_dvs[sheet_id] = dvs
    if all_dvs:
        import json
        resources.append({
            "name": "SHEET_DATA_VALIDATION_PLUGIN",
            "data": json.dumps(all_dvs, ensure_ascii=False),
        })

    # Pack conditional formatting (P5) into SHEET_CONDITIONAL_FORMATTING_PLUGIN.
    all_cf: dict[str, list[dict]] = {}
    for idx, ws in enumerate(wb.worksheets):
        sheet_id = f"sheet_{idx}"
        cf_rules = _read_conditional_formatting(ws)
        if cf_rules:
            all_cf[sheet_id] = cf_rules
    if all_cf:
        import json
        resources.append({
            "name": "SHEET_CONDITIONAL_FORMATTING_PLUGIN",
            "data": json.dumps(all_cf, ensure_ascii=False),
        })

    return {
        "id": "workbook",
        "name": "Workbook",
        "appVersion": "0.10.2",
        "locale": "zhCN",
        "styles": styles_registry.dump(),
        "sheetOrder": sheet_order,
        "sheets": sheets_univer,
        "resources": resources,
    }


# ── Floating image extraction (P4) ─────────────────────────────────────
# openpyxl exposes floating images via ws._images — each is an
# openpyxl.drawing.image.Image with:
#   .anchor: OneCellAnchor | TwoCellAnchor | AbsoluteAnchor
#   ._data(): bytes (method — call it)
#   .format: 'png' | 'jpeg' | 'gif' | ...
#
# Anchor structure:
#   OneCellAnchor:  .from + .ext (cx, cy in EMUs)  — position + size
#   TwoCellAnchor:  .from + .to                    — cell-bound, size implicit
#   AbsoluteAnchor: .pos (x, y in EMUs)            — pixel position (rare)
#
# EMU → px: 1 px = 9525 EMUs at 96 DPI.
# Univer transform.from/to use cell indices + cellOffset/rowOffset in px.

_EMU_PER_PX = 9525


def _read_drawings(ws, doc_id: str, sheet_id: str) -> tuple[dict[str, dict], int]:
    """Extract floating images from a worksheet → Univer drawing params.

    Returns (drawings_map, image_count). Each image is uploaded to MinIO
    under bucket 'collaboration' key 'docs/{doc_id}/images/{uuid}.{ext}'.
    """
    images = getattr(ws, "_images", None) or []
    if not images:
        return {}, 0

    from common import settings
    from common.misc_utils import get_uuid

    drawings: dict[str, dict] = {}
    count = 0
    for img in images:
        try:
            data_bytes = _extract_image_bytes(img)
            if not data_bytes:
                continue
            ext = (getattr(img, "format", None) or "png").lower()
            if ext == "jpeg":
                ext = "jpg"
            asset_id = get_uuid()
            storage_key = f"docs/{doc_id}/images/{asset_id}.{ext}"
            settings.STORAGE_IMPL.put("collaboration", storage_key, data_bytes)

            width, height = _image_natural_size(data_bytes, ext)
            transform = _anchor_to_transform(img, width, height)

            drawing_id = f"draw_{asset_id}"
            image_id = f"img_{asset_id}"
            url = (
                f"/api/v1/collaboration/documents/{doc_id}/assets/{asset_id}"
            )
            drawings[drawing_id] = _build_image_drawing(
                drawing_id=drawing_id,
                image_id=image_id,
                sheet_id=sheet_id,
                source=url,
                width=width,
                height=height,
                transform=transform,
                title="",
            )
            count += 1
        except Exception as e:
            logger.warning("skip image on sheet %s: %s", sheet_id, e)
            continue
    return drawings, count


def _extract_image_bytes(img) -> bytes | None:
    """Pull raw bytes from openpyxl Image. _data() is a method in newer versions."""
    try:
        data = img._data
        return data() if callable(data) else data
    except Exception:
        return None


# ── Chart placeholder (P6) ─────────────────────────────────────────────
# Free Univer has no chart plugin. To preserve user awareness of where
# charts existed in the source xlsx, we render a labeled PNG placeholder
# (chart type + series count) and embed it as a regular image drawing.
# Real chart round-trip is lost on import — this is a one-way preservation.

def _read_charts(ws, doc_id: str, sheet_id: str) -> tuple[dict[str, dict], int]:
    """Render each chart as a labeled PNG → upload → build drawing params."""
    charts = getattr(ws, "_charts", None) or []
    if not charts:
        return {}, 0

    from common import settings

    drawings: dict[str, dict] = {}
    count = 0
    for chart in charts:
        try:
            chart_kind = type(chart).__name__
            title = _extract_chart_title(chart)
            series_count = len(getattr(chart, "series", []) or [])

            png_bytes = _render_chart_placeholder(chart_kind, title, series_count)
            asset_id = f"chart_{_get_uuid()}"
            storage_key = f"docs/{doc_id}/images/{asset_id}.png"
            settings.STORAGE_IMPL.put("collaboration", storage_key, png_bytes)

            width, height = 480, 320
            anchor = getattr(chart, "anchor", None)
            transform = _chart_anchor_to_transform(anchor, width, height)

            drawing_id = f"draw_{asset_id}"
            image_id = f"img_{asset_id}"
            url = (
                f"/api/v1/collaboration/documents/{doc_id}/assets/{asset_id}"
            )
            drawings[drawing_id] = _build_image_drawing(
                drawing_id=drawing_id,
                image_id=image_id,
                sheet_id=sheet_id,
                source=url,
                width=width,
                height=height,
                transform=transform,
                title=f"[图表占位] {chart_kind}: {title}" if title else f"[图表占位] {chart_kind}",
            )
            count += 1
        except Exception as e:
            logger.warning("skip chart on sheet %s: %s", sheet_id, e)
            continue
    return drawings, count


def _extract_chart_title(chart) -> str:
    """Best-effort chart title extraction (openpyxl returns None or rich-text)."""
    title_obj = getattr(chart, "title", None)
    if title_obj is None:
        return ""
    # openpyxl Title has .tx.rich.p[0].r[0].t for rich text, or just str()
    try:
        if hasattr(title_obj, "tx") and title_obj.tx:
            rich = getattr(title_obj.tx, "rich", None)
            if rich and rich.p:
                runs = []
                for p in rich.p:
                    for r in (getattr(p, "r", []) or []):
                        runs.append(getattr(r, "t", "") or "")
                return "".join(runs)
        return str(title_obj) if title_obj else ""
    except Exception:
        return ""


def _render_chart_placeholder(chart_kind: str, title: str, series_count: int) -> bytes:
    """Generate a labeled PNG using PIL. ~480x320."""
    from PIL import Image, ImageDraw, ImageFont
    import io as _io

    W, H = 480, 320
    img = Image.new("RGB", (W, H), color=(250, 250, 250))
    draw = ImageDraw.Draw(img)
    # Border
    draw.rectangle([(0, 0), (W - 1, H - 1)], outline=(200, 200, 200), width=2)
    # Icon
    try:
        font_lg = ImageFont.truetype("arial.ttf", 28)
        font_sm = ImageFont.truetype("arial.ttf", 16)
    except OSError:
        font_lg = ImageFont.load_default()
        font_sm = ImageFont.load_default()
    # Chart emoji-like icon
    draw.text((W // 2 - 80, 100), "📊", fill=(100, 100, 100), font=font_lg)
    draw.text((W // 2 - 100, 150), f"图表占位 / Chart Placeholder",
              fill=(120, 120, 120), font=font_sm)
    draw.text((W // 2 - 80, 180), f"Type: {chart_kind}",
              fill=(80, 80, 80), font=font_sm)
    if title:
        draw.text((W // 2 - 100, 210), f"Title: {title[:30]}",
                  fill=(80, 80, 80), font=font_sm)
    draw.text((W // 2 - 100, 240), f"Series: {series_count}",
              fill=(80, 80, 80), font=font_sm)
    draw.text((W // 2 - 100, 280),
              "Univer 免费版图表已降级为占位图",
              fill=(150, 150, 150), font=font_sm)

    buf = _io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def _chart_anchor_to_transform(anchor, width: int, height: int) -> dict:
    """Charts use OneCellAnchor or TwoCellAnchor like images — reuse the image path."""
    # Build a fake-image wrapper so _anchor_to_transform can read its attrs
    class _ChartImgStub:
        pass
    stub = _ChartImgStub()
    stub.anchor = anchor
    return _anchor_to_transform(stub, width, height)


def _image_natural_size(data_bytes: bytes, ext: str) -> tuple[int, int]:
    """Determine image width/height in px. Falls back to (200, 150)."""
    try:
        from PIL import Image
        import io as _io
        with Image.open(_io.BytesIO(data_bytes)) as im:
            return im.size  # (w, h) in pixels
    except Exception:
        # No PIL or unsupported format — placeholder
        return 200, 150


def _anchor_to_transform(img, width: int, height: int) -> dict:
    """Convert openpyxl anchor → Univer sheet drawing position dicts.

    Returns { sheetTransform, axisAlignSheetTransform, transform }.
    - sheetTransform: { from:{column,columnOffset,row,rowOffset}, to:{...} }
    - transform: absolute { left, top, width, height, angle } — estimated
      using default cell sizes (Univer recomputes precisely on interaction).
    """
    anchor = getattr(img, "anchor", None)
    frm_col = frm_row = 0
    frm_col_off = frm_row_off = 0
    to_col = to_row = 0
    to_col_off = to_row_off = 0
    have_to = False

    if anchor is not None:
        # openpyxl exposes `xdr:from` as `_from` (Python keyword conflict).
        # Using `anchor.from` always returns None → all images collapse to (0,0).
        frm = getattr(anchor, "_from", None) or getattr(anchor, "from", None)
        to = getattr(anchor, "to", None)
        ext = getattr(anchor, "ext", None)
        if frm is not None:
            frm_col = frm.col or 0
            frm_row = frm.row or 0
            frm_col_off = int((frm.colOff or 0) / _EMU_PER_PX)
            frm_row_off = int((frm.rowOff or 0) / _EMU_PER_PX)
        if to is not None:
            to_col = to.col or 0
            to_row = to.row or 0
            to_col_off = int((to.colOff or 0) / _EMU_PER_PX)
            to_row_off = int((to.rowOff or 0) / _EMU_PER_PX)
            have_to = True
        elif ext is not None:
            ext_cx = getattr(ext, "cx", 0) or 0
            ext_cy = getattr(ext, "cy", 0) or 0
            w_px = int(ext_cx / _EMU_PER_PX) if ext_cx else width
            h_px = int(ext_cy / _EMU_PER_PX) if ext_cy else height
            to_col, to_row, to_col_off, to_row_off = _anchor_end_from_size(
                frm_col, frm_row, frm_col_off, frm_row_off, w_px, h_px
            )
            have_to = True

    if not have_to:
        to_col, to_row, to_col_off, to_row_off = _anchor_end_from_size(
            frm_col, frm_row, frm_col_off, frm_row_off, width, height
        )

    sheet_transform = {
        "from": {
            "column": frm_col,
            "columnOffset": frm_col_off,
            "row": frm_row,
            "rowOffset": frm_row_off,
        },
        "to": {
            "column": to_col,
            "columnOffset": to_col_off,
            "row": to_row,
            "rowOffset": to_row_off,
        },
    }
    # Estimate absolute position using default column width 100px / row 23px.
    # Univer recomputes precisely on first interaction via the sheet skeleton.
    abs_left = frm_col * 100 + frm_col_off
    abs_top = frm_row * 23 + frm_row_off
    abs_w = max(width, (to_col * 100 + to_col_off) - abs_left)
    abs_h = max(height, (to_row * 23 + to_row_off) - abs_top)
    absolute_transform = {
        "left": abs_left,
        "top": abs_top,
        "width": abs_w,
        "height": abs_h,
        "angle": 0,
        "flipX": False,
        "flipY": False,
    }
    return {
        "sheetTransform": sheet_transform,
        "axisAlignSheetTransform": sheet_transform,
        "transform": absolute_transform,
    }


def _anchor_end_from_size(frm_col, frm_row, frm_col_off, frm_row_off, w_px, h_px):
    """Compute a 'to' cell+offset that gives the desired pixel size."""
    start_x = frm_col * 100 + frm_col_off
    start_y = frm_row * 23 + frm_row_off
    end_x = start_x + w_px
    end_y = start_y + h_px
    to_col = end_x // 100
    to_row = end_y // 23
    to_col_off = end_x - to_col * 100
    to_row_off = end_y - to_row * 23
    return to_col, to_row, to_col_off, to_row_off


def _build_image_drawing(
    *,
    drawing_id: str,
    image_id: str,
    sheet_id: str,
    source: str,
    width: int,
    height: int,
    transform: dict,
    title: str,
) -> dict:
    """Build a Univer IImageData-compatible drawing dict.

    Univer's renderer (engine-render + sheets-drawing-ui) reads these fields
    DIRECTLY from the top-level drawing object:
      - imageSourceType / source  → image lookup
      - width / height            → image size
      - transform                 → absolute {left, top, width, height, angle}
      - sheetTransform             → cell-bound anchor for the sheet UI

    The nested {"image": {...}} shape used by an earlier adapter draft does NOT
    match what Univer's `drawingManagerData[unitId][subUnitId].data[drawingId]`
    expects and caused `Object.keys(null)` errors + non-rendering images.
    """
    return {
        "drawingId": drawing_id,
        # DrawingTypeEnum: DRAWING_IMAGE=0, DRAWING_SHAPE=1, DRAWING_CHART=2.
        # Previously was 1 (SHAPE) which silently failed to render — the image
        # renderer (c_e.renderImages) only fires when drawingType === DRAWING_IMAGE.
        "drawingType": 0,
        "unitId": "workbook",
        "subUnitId": sheet_id,
        "imageSourceType": "URL",
        "source": source,
        "imageId": image_id,
        "width": width,
        "height": height,
        "title": title,
        "sheetTransform": transform["sheetTransform"],
        "axisAlignSheetTransform": transform["axisAlignSheetTransform"],
        "transform": transform["transform"],
        # SheetDrawingAnchorType (Univer sheets-drawing):
        #   "0" = Position  — moves with cell insert/delete, size fixed (default)
        #   "1" = Both      — moves AND resizes with cell changes
        #   "2" = None      — fully fixed position/size
        # Without this field Univer defaults to "0", so images keep their
        # initial pixel size even when the user resizes the anchor columns or
        # rows. We emit "1" so twoCellAnchor images stretch with the cell range,
        # matching Excel's default twoCell behavior.
        "anchorType": "1",
    }


# ── Cell notes (P5) ────────────────────────────────────────────────────
# openpyxl: cell.comment is a CellComment with .text and .author.
# Univer ISheetNote resource: { sheetId: { row: { col: ISheetNote } } }

from common.misc_utils import get_uuid as _get_uuid  # noqa: E402


def _read_notes(ws) -> dict[str, dict[str, dict]]:
    """Walk cells with comments → Univer note resource shape."""
    out: dict[str, dict[str, dict]] = {}
    for row in ws.iter_rows():
        for cell in row:
            comment = getattr(cell, "comment", None)
            if comment is None:
                continue
            try:
                row_idx = cell.row - 1
                col_idx = cell.column - 1
            except (AttributeError, TypeError):
                continue
            text = getattr(comment, "text", None) or ""
            note_obj = {
                "id": f"note_{_get_uuid()}",
                "row": row_idx,
                "col": col_idx,
                "width": 120,
                "height": 80,
                "note": text,
                "show": False,
            }
            out.setdefault(str(row_idx), {})[str(col_idx)] = note_obj
    return out


# ── Data validation (P5) ───────────────────────────────────────────────
# openpyxl ws.data_validations.dataValidation: list of DataValidation
#   .type: 'list' | 'whole' | 'decimal' | 'date' | 'time' | 'textLength' | 'custom'
#   .formula1: e.g. '"a,b,c"' for list, '0' for whole lower bound
#   .formula2: upper bound (for between)
#   .operator: 'between' | 'notBetween' | 'equal' | 'notEqual' | 'greaterThan' | ...
#   .sqref: cell range string e.g. "A1:A10"
# Univer resource shape: { sheetId: [IDataValidation] }
# IDataValidation: { uid, type, formula1, formula2, operator, ranges: [{startRow,...}] }

_DV_TYPE_MAP = {
    "list": "list",
    "whole": "whole",
    "decimal": "decimal",
    "date": "date",
    "time": "time",
    "textLength": "textLength",
    "custom": "custom",
}


def _read_data_validations(ws) -> list[dict]:
    dvs: list[dict] = []
    dv_collection = getattr(ws, "data_validations", None)
    if not dv_collection:
        return dvs
    for dv in getattr(dv_collection, "dataValidation", []) or []:
        try:
            ranges = _parse_sqref(getattr(dv, "sqref", "") or "")
            if not ranges:
                continue
            dv_type = _DV_TYPE_MAP.get(getattr(dv, "type", "") or "")
            if not dv_type:
                continue
            entry: dict[str, Any] = {
                "uid": f"dv_{_get_uuid()}",
                "type": dv_type,
                "formula1": getattr(dv, "formula1", None) or "",
                "formula2": getattr(dv, "formula2", None) or "",
                "operator": getattr(dv, "operator", "between") or "between",
                "ranges": ranges,
                "allowBlank": bool(getattr(dv, "allowBlank", True)),
                "showInputMessage": bool(getattr(dv, "showInputMessage", False)),
                "showErrorMessage": bool(getattr(dv, "showErrorMessage", False)),
                "errorStyle": "stop",
                "renderMode": "arrow" if dv_type == "list" else "square",
            }
            # For list type, formula1 is a quoted comma string: '"a,b,c"' or range ref
            f1 = entry["formula1"]
            if dv_type == "list" and f1.startswith('"') and f1.endswith('"'):
                entry["formula1"] = f1[1:-1]
            dvs.append(entry)
        except Exception as e:
            logger.warning("skip data validation: %s", e)
    return dvs


def _parse_sqref(sqref) -> list[dict]:
    """Convert Excel sqref (e.g. 'A1:A10' or 'B2 D5:E6') → Univer ranges."""
    from openpyxl.utils.cell import range_boundaries
    if not sqref:
        return []
    # sqref can be a MultiCellRange or a string
    ranges_str = str(sqref)
    out: list[dict] = []
    for part in ranges_str.split():
        try:
            min_col, min_row, max_col, max_row = range_boundaries(part)
            out.append({
                "startRow": min_row - 1,
                "endRow": max_row - 1,
                "startColumn": min_col - 1,
                "endColumn": max_col - 1,
            })
        except Exception:
            continue
    return out


# ── Conditional formatting (P5) ────────────────────────────────────────
# openpyxl ws.conditional_formatting yields (cell_range, rules) tuples.
# Each rule has .type ('cellIs', 'colorScale', 'dataBar', 'expression', ...)
# Univer IConditionFormattingRule:
#   { cfId, ranges: [{startRow,...}], stopIfTrue: bool, rule: { type, ... } }

def _read_conditional_formatting(ws) -> list[dict]:
    rules_out: list[dict] = []
    cf_collection = getattr(ws, "conditional_formatting", None)
    if not cf_collection:
        return rules_out
    try:
        # openpyxl ConditionalFormattingList: iterate yields ConditionalFormatting
        # objects with .sqref (MultiCellRange) and .rules (list of Rule).
        for cf in cf_collection:
            ranges = _parse_cf_range(cf)
            if not ranges:
                continue
            for rule in getattr(cf, "rules", []) or []:
                univer_rule = _convert_cf_rule(rule)
                if not univer_rule:
                    continue
                rules_out.append({
                    "cfId": f"cf_{_get_uuid()}",
                    "ranges": ranges,
                    "stopIfTrue": bool(getattr(rule, "stopIfTrue", False)),
                    "rule": univer_rule,
                })
    except Exception as e:
        logger.warning("conditional formatting read failed on sheet %s: %s", ws.title, e)
    return rules_out


def _parse_cf_range(cf_range) -> list[dict]:
    """ConditionalFormatting sqref → Univer ranges list."""
    from openpyxl.utils.cell import range_boundaries
    ranges: list[dict] = []
    try:
        sqref = str(getattr(cf_range, "sqref", "") or cf_range)
    except Exception:
        sqref = str(cf_range)
    for part in sqref.split():
        try:
            min_col, min_row, max_col, max_row = range_boundaries(part)
            ranges.append({
                "startRow": min_row - 1,
                "endRow": max_row - 1,
                "startColumn": min_col - 1,
                "endColumn": max_col - 1,
            })
        except Exception:
            continue
    return ranges


def _convert_cf_rule(rule) -> dict | None:
    """Map openpyxl CF rule type → Univer CF rule config.

    Handled: cellIs, colorScale, dataBar, expression, containsText, timePeriod.
    Returns None for unsupported types so the caller can skip them.
    """
    rule_type = getattr(rule, "type", None) or ""
    style = _extract_cf_style(rule)

    if rule_type == "cellIs":
        operator = _map_cf_operator(getattr(rule, "operator", None))
        formula = list(getattr(rule, "formula", []) or [])
        # Univer INumberHighlightCell requires numeric value(s).
        # between / notBetween need [num1, num2]; others need single number.
        nums: list[float] = []
        for raw in formula:
            try:
                nums.append(float(raw))
            except (TypeError, ValueError):
                # non-numeric formula (e.g. a cell reference) — skip
                continue
        if operator in ("between", "notBetween"):
            if len(nums) >= 2:
                value: float | list[float] = [nums[0], nums[1]]
            elif nums:
                value = [nums[0], nums[0]]
            else:
                value = [0, 0]
        else:
            value = nums[0] if nums else 0
        return {
            "type": "highlightCell",
            "subType": "number",
            "operator": operator,
            "value": value,
            "style": style,
        }
    if rule_type == "colorScale":
        return {
            "type": "colorScale",
            "config": _convert_color_scale(rule),
        }
    if rule_type == "dataBar":
        return {
            "type": "dataBar",
            "isShowValue": True,
            "config": _convert_data_bar(rule),
        }
    if rule_type in ("expression", "containsText", "notContainsText", "beginsWith", "endsWith"):
        sub_map = {
            "expression": "formula",
            "containsText": "text",
            "notContainsText": "text",
            "beginsWith": "text",
            "endsWith": "text",
        }
        operator_map = {
            "containsText": "containsText",
            "notContainsText": "notContainsText",
            "beginsWith": "beginsWith",
            "endsWith": "endsWith",
        }
        formula = list(getattr(rule, "formula", []) or [])
        return {
            "type": "highlightCell",
            "subType": sub_map.get(rule_type, "formula"),
            "operator": operator_map.get(rule_type, "containsText"),
            "value": formula[0] if formula else "",
            "style": style,
        }
    if rule_type == "timePeriod":
        formula = list(getattr(rule, "formula", []) or [])
        return {
            "type": "highlightCell",
            "subType": "timePeriod",
            "operator": getattr(rule, "timePeriod", "yesterday") or "yesterday",
            "style": style,
        }
    # Unsupported rule types: aboveAverage, top10, uniqueValues, iconSet —
    # log and skip rather than emit a malformed rule.
    logger.debug("skip unsupported CF rule type: %s", rule_type)
    return None


def _map_cf_operator(opx_op) -> str:
    return {
        "between": "between",
        "notBetween": "notBetween",
        "equal": "equal",
        "notEqual": "notEqual",
        "greaterThan": "greaterThan",
        "greaterThanOrEqual": "greaterThanOrEqual",
        "lessThan": "lessThan",
        "lessThanOrEqual": "lessThanOrEqual",
    }.get(opx_op or "", "between")


def _extract_cf_style(rule) -> dict:
    """Extract Univer style dict from openpyxl DifferentialStyle."""
    dxf = getattr(rule, "dxf", None)
    if not dxf:
        return {}
    style: dict[str, Any] = {}
    font = getattr(dxf, "font", None)
    if font:
        if font.bold:
            style["bl"] = 1
        if font.italic:
            style["it"] = 1
        if font.color is not None:
            rgb = _argb_to_rgb(font.color)
            if rgb:
                style["cl"] = {"rgb": rgb}
    fill = getattr(dxf, "fill", None)
    if fill and fill.fill_type == "solid":
        rgb = _argb_to_rgb(getattr(fill, "fgColor", None))
        if rgb:
            style["bg"] = {"rgb": rgb}
    return style


def _convert_color_scale(rule) -> list:
    """openpyxl ColorScaleRule → Univer colorScale config array."""
    config: list[dict] = []
    colors = list(getattr(rule, "colors", []) or [])
    values = list(getattr(rule, "values", []) or [])
    types = list(getattr(rule, "types", []) or [])
    for idx in range(max(len(colors), len(values), 3)):
        try:
            color = colors[idx] if idx < len(colors) else None
            value = values[idx] if idx < len(values) else None
            vtype = types[idx] if idx < len(types) else "num"
            rgb = _argb_to_rgb(color) if color is not None else "#FFFFFF"
            config.append({
                "index": idx,
                "color": rgb or "#FFFFFF",
                "value": {"type": vtype, "value": value if value is not None else 0},
            })
        except Exception:
            continue
    return config


def _convert_data_bar(rule) -> dict:
    """openpyxl DataBarRule → Univer dataBar config."""
    return {
        "min": {"type": "auto"},
        "max": {"type": "auto"},
        "isGradient": True,
        "positiveColor": "#638EC6",
        "nativeColor": "#F8696B",
    }


def _worksheet_to_univer(ws, sheet_id: str, idx: int, styles_registry: "_StyleRegistry") -> dict:
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row > MAX_ROWS:
        raise XlsxTooLargeError(
            f"行数过多：{ws.title} 有 {max_row} 行（上限 {MAX_ROWS}）",
            code="TOO_MANY_ROWS",
        )
    if max_col > MAX_COLS:
        raise XlsxTooLargeError(
            f"列数过多：{ws.title} 有 {max_col} 列（上限 {MAX_COLS}）",
            code="TOO_MANY_COLS",
        )

    cell_data: dict[str, dict[str, dict]] = {}
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col)
    ):
        for col_idx, cell in enumerate(row):
            record = _cell_to_univer(cell, styles_registry)
            if record is None:
                continue
            r_key = str(row_idx)
            c_key = str(col_idx)
            cell_data.setdefault(r_key, {})[c_key] = record

    column_data = _read_column_dimensions(ws)
    row_data = _read_row_dimensions(ws)
    merge_data = _read_merges(ws)
    freeze = _read_freeze(ws)

    return {
        "id": sheet_id,
        "name": ws.title or f"Sheet{idx + 1}",
        "tabColor": "",
        "hidden": 1 if ws.sheet_state == "hidden" else 0,
        "rowCount": max(max_row + 100, 200),
        "columnCount": max(max_col + 20, 26),
        "zoomRatio": 1,
        "scrollTop": 0,
        "scrollLeft": 0,
        "defaultColumnWidth": 100,
        "defaultRowHeight": 23,
        "freeze": freeze,
        "mergeData": merge_data,
        "cellData": cell_data,
        "rowData": row_data,
        "columnData": column_data,
        "showGridlines": 1,
        "rowHeader": {"width": 46},
        "columnHeader": {"height": 20},
        "rightToLeft": 1 if getattr(ws.sheet_view, "rightToLeft", False) else 0,
    }


def _cell_to_univer(cell, styles_registry: "_StyleRegistry") -> dict | None:
    v = cell.value
    is_empty = v is None

    # Build style first — even empty cells can carry style (e.g. colored blanks
    # from a formatted template). If neither value nor style, skip entirely.
    style_id = _build_style(cell, styles_registry)
    if is_empty and style_id is None:
        return None

    record: dict[str, Any] = {}
    if style_id is not None:
        record["s"] = style_id
    if is_empty:
        # Empty cell with style — return a minimal record so the style is honored
        return record

    # Formula — openpyxl returns formulas as strings starting with '='
    if isinstance(v, str) and v.startswith("="):
        record["f"] = v[1:]
        record["t"] = 1
        return record

    # Number (int / float — but bool is also int subclass, check first)
    if isinstance(v, bool):
        record["v"] = v
        record["t"] = 3
        return record
    if isinstance(v, (int, float)):
        record["v"] = v
        record["t"] = 2
        return record

    # Datetime / date / time / timedelta → ISO string
    if hasattr(v, "isoformat"):
        try:
            record["v"] = v.isoformat()
            record["t"] = 1
            return record
        except Exception:
            pass

    record["v"] = str(v)
    record["t"] = 1
    return record


# ── Style conversion (P3) ──────────────────────────────────────────────
# openpyxl: cell.font / cell.fill / cell.border / cell.alignment / cell.number_format
# Univer IStyleData shape:
#   {
#     ff: font family, fs: font size, bl: bold(1/0), it: italic,
#     ul: {s:1}, st: {s:1}, cl: {rgb}, bg: {rgb},
#     bd: {t,b,l,r: {s: borderStyle, c: {rgb}}},
#     ht: horizontal align (1=left,2=center,3=right),
#     vt: vertical align   (1=middle,2=top,3=bottom),
#     tb: wrap text (1/0),
#     n:  number format pattern,
#   }

# Univer border style enum strings → openpyxl Side.style strings map 1:1 for the common cases
_BORDER_STYLE_MAP = {
    "thin": "thin",
    "medium": "medium",
    "thick": "thick",
    "dotted": "dotted",
    "dashed": "dashed",
    "double": "double",
    "dashDot": "dashDot",
    "dashDotDot": "dashDotDot",
    "slantDashDot": "slantDashDot",
    "mediumDashed": "mediumDashed",
    "mediumDashDot": "mediumDashDot",
    "mediumDashDotDot": "mediumDashDotDot",
}

_HORIZONTAL_ALIGN_MAP = {
    "left": 1,
    "center": 2,
    "right": 3,
    "justify": 4,  # Univer supports distributed/justify via different codes; 4 is acceptable
    "distributed": 7,
}

_VERTICAL_ALIGN_MAP = {
    "top": 2,
    "center": 1,  # Univer uses 1 for middle (vertical center)
    "bottom": 3,
    "justify": 4,
    "distributed": 7,
}


class _StyleRegistry:
    """Dedup IStyleData objects via JSON-string hash. Returns integer IDs as strings.

    Univer's workbook.styles is keyed by string; building many duplicate style
    dicts (e.g. every header cell) would balloon the JSON. Hash + dedup keeps
    a 10k-row sheet with 5 distinct styles down to 5 entries.
    """

    def __init__(self) -> None:
        self._by_hash: dict[str, str] = {}
        self._by_id: dict[str, dict] = {}

    def register(self, style: dict) -> str | None:
        """Return stable style ID for the given style dict. None if empty."""
        if not style:
            return None
        key = _json_dumps_sorted(style)
        existing = self._by_hash.get(key)
        if existing is not None:
            return existing
        sid = str(len(self._by_hash))
        self._by_hash[key] = sid
        self._by_id[sid] = style
        return sid

    def dump(self) -> dict[str, dict]:
        return dict(self._by_id)


def _json_dumps_sorted(obj: Any) -> str:
    import json
    return json.dumps(obj, sort_keys=True, separators=(",", ":"))


def _build_style(cell, registry: _StyleRegistry) -> str | None:
    """Extract style attributes from openpyxl cell and register, return style ID."""
    style: dict[str, Any] = {}

    # ── Font ──
    font = getattr(cell, "font", None)
    if font and font.name:
        style["ff"] = font.name
    if font and font.size:
        style["fs"] = font.size
    if font and font.bold:
        style["bl"] = 1
    if font and font.italic:
        style["it"] = 1
    if font and font.underline and font.underline != "none":
        # Univer underline: { s: 1, t?: 'single'|'double' }
        ut = "double" if font.underline == "double" else "single"
        style["ul"] = {"s": 1, "t": ut}
    if font and font.strike:
        style["st"] = {"s": 1}
    if font and font.color is not None:
        rgb = _argb_to_rgb(font.color)
        if rgb:
            style["cl"] = {"rgb": rgb}

    # ── Fill / background ──
    fill = getattr(cell, "fill", None)
    if fill and fill.fill_type == "solid":
        # PatternFill uses fgColor as the visible color for "solid"
        rgb = _argb_to_rgb(getattr(fill, "fgColor", None))
        if rgb:
            style["bg"] = {"rgb": rgb}

    # ── Borders ──
    border = getattr(cell, "border", None)
    if border:
        bd: dict[str, Any] = {}
        for side_name, side_obj in (
            ("t", getattr(border, "top", None)),
            ("b", getattr(border, "bottom", None)),
            ("l", getattr(border, "left", None)),
            ("r", getattr(border, "right", None)),
        ):
            if not side_obj or not side_obj.style:
                continue
            univer_style = _BORDER_STYLE_MAP.get(side_obj.style, side_obj.style)
            entry: dict[str, Any] = {"s": univer_style}
            rgb = _argb_to_rgb(getattr(side_obj, "color", None))
            if rgb:
                entry["c"] = {"rgb": rgb}
            bd[side_name] = entry
        if bd:
            style["bd"] = bd

    # ── Alignment ──
    align = getattr(cell, "alignment", None)
    if align:
        if align.horizontal and align.horizontal in _HORIZONTAL_ALIGN_MAP:
            style["ht"] = _HORIZONTAL_ALIGN_MAP[align.horizontal]
        if align.vertical and align.vertical in _VERTICAL_ALIGN_MAP:
            style["vt"] = _VERTICAL_ALIGN_MAP[align.vertical]
        if align.wrap_text:
            style["tb"] = 1
        if align.text_rotation:
            # Univer uses degrees (0-180). openpyxl also uses degrees.
            style["rt"] = int(align.text_rotation)

    # ── Number format ──
    # openpyxl exposes the format string directly; Univer's `n` field can be a
    # pattern string. Skip the default "General" since it's a no-op.
    num_fmt = getattr(cell, "number_format", None)
    if num_fmt and num_fmt and num_fmt not in ("General", "general"):
        style["n"] = num_fmt

    return registry.register(style)


def _argb_to_rgb(color) -> str | None:
    """Convert openpyxl Color → Univer-style "#RRGGBB" hex.

    Univer's `IColorData.rgb` expects a leading `#` (e.g. `"#FF0000"`).
    Without it, Univer fails to parse the color and falls back to black.

    openpyxl colors come in three flavors:
      1. .rgb = "FFRRGGBB" (aRGB with alpha) — most common, strip leading FF
      2. .theme = integer (theme color) — needs workbook theme lookup (P3 MVP skips)
      3. .indexed = integer (legacy indexed palette) — use STAFF_PALETTE lookup
    """
    if color is None:
        return None

    # RGB hex (preferred)
    rgb = getattr(color, "rgb", None)
    if isinstance(rgb, str) and len(rgb) == 8:
        return "#" + rgb[2:].upper()  # strip alpha, add # for Univer
    if isinstance(rgb, str) and len(rgb) == 6:
        return "#" + rgb.upper()

    # Indexed palette (legacy xlsx)
    indexed = getattr(color, "indexed", None)
    if isinstance(indexed, int):
        # openpyxl's COLOR_INDEX has 64 entries; default Excel palette.
        # The well-known first 16 are usable for legacy files.
        palette = _LEGACY_PALETTE
        if 0 <= indexed < len(palette):
            return "#" + palette[indexed]

    # Theme color — would require parsing workbook theme XML.
    # P3 MVP: return None so the cell renders with default color.
    # Future: load workbook theme → resolve theme+tint → RGB.
    return None


# Standard Excel 64-color legacy palette (subset). Used only for old xlsx files
# that use indexed colors. Values from openpyxl's COLOR_INDEX first 16.
_LEGACY_PALETTE = [
    "000000", "FFFFFF", "FF0000", "00FF00", "0000FF", "FFFF00", "FF00FF", "00FFFF",
    "000000", "FFFFFF", "FF0000", "00FF00", "0000FF", "FFFF00", "FF00FF", "00FFFF",
]


def _read_column_dimensions(ws) -> dict[int, dict]:
    column_data: dict[int, dict] = {}
    for col_letter, dim in ws.column_dimensions.items():
        try:
            col_idx = _col_letter_to_idx(col_letter)
        except ValueError:
            continue
        entry: dict[str, Any] = {}
        if dim.width:
            entry["w"] = max(_MIN_COL_W, int(dim.width * _PX_PER_CHAR) + _COL_PADDING)
        if dim.hidden:
            entry["hd"] = 1
        if entry:
            column_data[col_idx] = entry
    return column_data


def _read_row_dimensions(ws) -> dict[int, dict]:
    row_data: dict[int, dict] = {}
    for row_num, dim in ws.row_dimensions.items():
        row_idx = row_num - 1  # 1-indexed → 0-indexed
        entry: dict[str, Any] = {}
        if dim.height:
            entry["h"] = max(_MIN_ROW_H, int(dim.height * _PX_PER_POINT))
            # CRITICAL: mark height as explicit (not auto). Univer's default for
            # `ha` is true (auto-height), which means a column-width resize
            # triggers content-based recalculation and discards the imported `h`.
            # Symptom without this flag: drag a column border → all custom row
            # heights reset to default text height. Dragging row height directly
            # works only because Univer's UI sets `ha: false` on user drag.
            entry["ha"] = False
        if dim.hidden:
            entry["hd"] = 1
        if entry:
            row_data[row_idx] = entry
    return row_data


def _read_merges(ws) -> list[dict]:
    merges: list[dict] = []
    for r in ws.merged_cells.ranges:
        merges.append({
            "startRow": r.min_row - 1,
            "endRow": r.max_row - 1,
            "startColumn": r.min_col - 1,
            "endColumn": r.max_col - 1,
        })
    return merges


def _read_freeze(ws) -> dict:
    freeze = {"startRow": -1, "startColumn": -1, "ySplit": 0, "xSplit": 0}
    fp = ws.freeze_panes
    if not fp:
        return freeze
    col_letter = "".join(c for c in str(fp) if c.isalpha())
    row_str = "".join(c for c in str(fp) if c.isdigit())
    if not col_letter and not row_str:
        return freeze
    col_idx = _col_letter_to_idx(col_letter) if col_letter else 0
    row_idx = (int(row_str) - 1) if row_str else 0
    return {
        "startRow": row_idx,
        "startColumn": col_idx,
        "ySplit": row_idx,
        "xSplit": col_idx,
    }


def _col_letter_to_idx(letter: str) -> int:
    """A→0, B→1, ..., Z→25, AA→26, ..."""
    if not letter:
        raise ValueError("empty column letter")
    result = 0
    for c in letter.upper():
        if not ("A" <= c <= "Z"):
            raise ValueError(f"bad column letter: {letter!r}")
        result = result * 26 + (ord(c) - ord("A") + 1)
    return result - 1


def _blank_sheet(sheet_id: str, name: str) -> dict:
    return {
        "id": sheet_id,
        "name": name,
        "tabColor": "",
        "hidden": 0,
        "rowCount": 50,
        "columnCount": 26,
        "zoomRatio": 1,
        "scrollTop": 0,
        "scrollLeft": 0,
        "defaultColumnWidth": 100,
        "defaultRowHeight": 23,
        "freeze": {"startRow": -1, "startColumn": -1, "ySplit": 0, "xSplit": 0},
        "mergeData": [],
        "cellData": {},
        "rowData": {},
        "columnData": {},
        "showGridlines": 1,
        "rowHeader": {"width": 46},
        "columnHeader": {"height": 20},
        "rightToLeft": 0,
    }


# ── Univer → openpyxl ──────────────────────────────────────────────────

def _write_legacy_sheets(wb: Workbook, sheets: list) -> None:
    """Backward compat — legacy {sheets:[{name,data,colWidths}]} format."""
    for i, sheet_def in enumerate(sheets):
        name = sheet_def.get("name", f"Sheet{i + 1}")
        ws = wb.active if i == 0 else wb.create_sheet(title=name)
        if i == 0:
            ws.title = name

        for r_idx, row in enumerate(sheet_def.get("data", [])):
            for c_idx, val in enumerate(row):
                if val is None or val == "":
                    continue
                if isinstance(val, str):
                    converted: Any = val
                    try:
                        converted = int(val)
                    except ValueError:
                        try:
                            converted = float(val)
                        except ValueError:
                            pass
                    val = converted
                ws.cell(row=r_idx + 1, column=c_idx + 1, value=val)

        for c_idx, width in enumerate(sheet_def.get("colWidths", [])):
            ws.column_dimensions[get_column_letter(c_idx + 1)].width = max(
                8, (width or 100) / _PX_PER_CHAR
            )


def _write_univer_sheets(wb: Workbook, sheets: dict, sheet_order: list, styles: dict | None = None) -> None:
    sheet_ids = sheet_order or list(sheets.keys())
    styles = styles or {}
    for i, sheet_id in enumerate(sheet_ids):
        sheet = sheets.get(sheet_id) or {}
        name = sheet.get("name") or f"Sheet{i + 1}"
        ws = wb.active if i == 0 else wb.create_sheet(title=name)
        if i == 0:
            ws.title = name

        # Cells
        for row_str, row_cells in (sheet.get("cellData") or {}).items():
            try:
                row_idx = int(row_str)
            except ValueError:
                continue
            for col_str, cell_record in row_cells.items():
                try:
                    col_idx = int(col_str)
                except ValueError:
                    continue
                value = _univer_cell_to_excel(cell_record)
                style_id = cell_record.get("s") if isinstance(cell_record, dict) else None
                # Skip only when there's no value AND no style — otherwise
                # empty-but-styled cells would lose their background/border.
                if value is None and style_id is None:
                    continue
                cell = ws.cell(row=row_idx + 1, column=col_idx + 1, value=value)
                # Apply style if referenced
                if style_id is not None:
                    style = styles.get(str(style_id))
                    if style:
                        _apply_style_to_cell(cell, style)

        # Column widths / hidden
        for col_str, col_def in (sheet.get("columnData") or {}).items():
            try:
                col_idx = int(col_str)
            except ValueError:
                continue
            letter = get_column_letter(col_idx + 1)
            dim = ws.column_dimensions[letter]
            w = col_def.get("w")
            if w:
                dim.width = max(8, (w - _COL_PADDING) / _PX_PER_CHAR)
            if col_def.get("hd"):
                dim.hidden = True

        # Row heights / hidden
        for row_str, row_def in (sheet.get("rowData") or {}).items():
            try:
                row_num = int(row_str) + 1
            except ValueError:
                continue
            dim = ws.row_dimensions[row_num]
            h = row_def.get("h")
            if h:
                dim.height = max(15, h / _PX_PER_POINT)
            if row_def.get("hd"):
                dim.hidden = True

        # Merged cells
        for merge in sheet.get("mergeData") or []:
            try:
                start = (
                    get_column_letter(merge["startColumn"] + 1)
                    + str(merge["startRow"] + 1)
                )
                end = (
                    get_column_letter(merge["endColumn"] + 1)
                    + str(merge["endRow"] + 1)
                )
                ws.merge_cells(f"{start}:{end}")
            except (KeyError, Exception) as e:
                logger.warning("skip bad merge entry %r: %s", merge, e)

        # Freeze panes
        freeze = sheet.get("freeze") or {}
        start_row = freeze.get("startRow", -1)
        start_col = freeze.get("startColumn", -1)
        if start_row >= 0 or start_col >= 0:
            row_part = (start_row + 1) if start_row >= 0 else 1
            col_idx = (start_col + 1) if start_col >= 0 else 1
            col_part = get_column_letter(max(col_idx, 1))
            ws.freeze_panes = f"{col_part}{row_part}"

        # Hidden sheet
        if sheet.get("hidden"):
            ws.sheet_state = "hidden"

        # RTL
        if sheet.get("rightToLeft"):
            try:
                ws.sheet_view.rightToLeft = True
            except Exception:
                pass


def _univer_cell_to_excel(record: dict):
    """Convert a Univer ICellData → openpyxl-compatible value."""
    if not isinstance(record, dict):
        return None
    # Formula takes precedence
    f = record.get("f")
    if isinstance(f, str) and f:
        return "=" + f
    v = record.get("v")
    if v is None:
        return None
    t = record.get("t", 1)
    if t == 2:  # number
        if isinstance(v, (int, float)):
            return v
        try:
            return float(v) if "." in str(v) else int(v)
        except (TypeError, ValueError):
            return v
    if t == 3:  # bool
        if isinstance(v, bool):
            return v
        if isinstance(v, str):
            return v.strip().lower() in ("true", "1", "yes")
        return bool(v)
    # String / default
    return v
