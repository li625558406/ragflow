"""
fujian_jtyst_zdgksxml_crawler — 福建省交通运输-公开事项 智能采集自定义 runner
===========================================================================

数据源变更说明 (2026-07-28):
    旧版本: 抓取 matterData/{id}.json 嵌套树，82 个叶子，无超链接。
    新版本: 抓取"目录下载"按钮指向的 xlsx 表格 (211 行 × 13 列)，col F 内嵌
            209 个超链接指向国家法规原文。粒度从"叶子节点"细化到"(公开事项, URL)"
            组合，约 184 条；同 URL 被多个事项引用时各为独立条目（result_id 由
            source_url + 锚点区分）。

数据流:
    1. 删除本地旧 xlsx（若存在）→ 重新下载 matterData/{id}.xlsx
    2. openpyxl 解析: 构建合并单元格 row→representative 映射，遍历 R3..Rmax
       继承 category_l1/l2/matter；每个 (matter, url) 收集多条条款行合并
    3. 对每个 item:
       - source_url = legal_doc_url + "#matter-{seqN|matter_hash}"  (仅用于 PK 唯一)
         锚点 token 取稳定值（xlsx 序号列 seq_no 优先；缺失时退到 matter 的
         md5 短 hash；最后兜底 idx），避免 xlsx 行序变化导致 PK 漂移。
       - legal_doc_url = xlsx col F hyperlink (真实抓取目标)
       - news_type = "福建省交通运输-公开事项" (固定)
       - content   = 13 列结构化字段拼成的 markdown + 详情页正文（HTML→MD）
       - 结构化字段写入 CollectionZdgksxmlExt 扩展表
    4. StoragePipeline.store(NormalizedItem):
       - CollectionWriter 写 crawler_result + collection_zdgksxml_ext
       - KBUploader 上传 markdown 到 KB
    5. summary 回传，crawler_task.last_run_* 由 unified_crawler 回写

去重 / 增量:
    crawler_result PK = md5(site_id|source_url)。
    source_url 含 #matter-{seq_no} 锚点，使每个 (matter, url) 都有稳定独立 PK。
    重复触发 = upsert，新增行 results_new + 1，已存在行 results_updated + 1。
    xlsx 行序变化（增删/重排）不会引起 PK 漂移（锚点基于内容稳定值）。
    每次运行先删本地 xlsx 再重下载，确保使用最新目录。

详情抓取容错:
    - 选择器优先级: .TRS_Editor → .pages_content → .content-text → #zoom →
      .article-content → article → .content → .detail → 最大 <div> → 全 <p>
    - 失败兜底: 仅用 xlsx 13 列拼装 content，不阻断流程
    - 节流 0.5-1.5s, timeout=20s, verify_ssl=False

调用入口:
    由 unified_crawler.py 的 custom_runner 分支调度（site_config.custom_runner
    字段指向本模块）。

参考文档:
    智能采集系统设计.md §4.4.1 / §4.5 SOP
"""
import hashlib
import logging
import os
import random
import re
import sys
import time
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import urlparse

import requests

# 项目路径初始化（容器内 /ragflow，本地 dev 时手动 cwd）
_THIS_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.abspath(os.path.join(_THIS_DIR, "..", ".."))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

from rag.svr.crawler_engine.models import item_from_dict

# ── 常量 ──────────────────────────────────────────────────────────────
SITE_ID = "fujian_jtyst_zdgksxml"
SITE_NAME = "福建省交通运输-公开事项"
SITE_DOMAIN = "jtyst.fujian.gov.cn"
NEWS_TYPE = "福建省交通运输-公开事项"  # 用户需求：列表类型固定
SOURCE_NAME = "福建省交通运输厅"
TARGET_PAGE_URL = "https://jtyst.fujian.gov.cn/zwgk/zfxxgkzl/zdgksxml/"
XLSX_URL = "https://jtyst.fujian.gov.cn/matterData/4028918175536a3b0175536f92043058.xlsx"
XLSX_FILENAME = "catalog.xlsx"

DEFAULT_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
              "application/octet-stream,*/*",
    "Accept-Language": "zh-CN,zh;q=0.9",
    "Referer": TARGET_PAGE_URL,
}

DETAIL_HEADERS = {
    **DEFAULT_HEADERS,
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}
DETAIL_TIMEOUT = 20
DETAIL_CONTENT_MIN_LEN = 80  # 抓取正文少于这个值视为空

REQUEST_DELAY_MIN = 0.5
REQUEST_DELAY_MAX = 1.5

# 详情页正文 CSS 选择器优先级（覆盖 TRS CMS / gov.cn / 国务院政策库 / moj.gov.cn 等）
DETAIL_SELECTORS: Tuple[str, ...] = (
    ".TRS_Editor",       # fujian.gov.cn TRS CMS
    ".pages_content",    # www.gov.cn 政策文件
    ".content-text",     # xzfg.moj.gov.cn 国家行政法规库
    "#zoom",             # 部分人民网/新华网转载
    ".article-content",
    ".article_content",
    "article",
    ".content",
    ".detail-content",
    ".detail",
)

# xlsx 表头 → 标准字段名映射（R2 表头行）
_HEADER_MAP = {
    "序号": "seq_no",
    "公开类别": "category_l1",
    "二级公开类别": "category_l2",
    "二级类别": "category_l2",
    "公开事项": "matter",
    "公开内容": "disclosure_content",
    "公开依据文件": "legal_doc_title",
    "公开依据文件条款": "legal_doc_clause",
    "公开时限": "disclosure_deadline",
    "公开期限": "disclosure_period",
    "公开主体": "disclosure_subject",
    "公开责任": "disclosure_duty",
    "公开方式": "disclosure_method",
    "公开渠道": "disclosure_channel",
}

logger = logging.getLogger(__name__)


# ── HTTP helpers ──────────────────────────────────────────────────────
def _throttle() -> None:
    time.sleep(random.uniform(REQUEST_DELAY_MIN, REQUEST_DELAY_MAX))


def _http_get(url: str, headers: Optional[Dict[str, str]] = None,
              timeout: int = 20, as_bytes: bool = False) -> requests.Response:
    """GET with retry (1 attempt, callers handle exceptions).

    verify_ssl=False — 政府站点常见自签名/过期证书。
    """
    resp = requests.get(
        url,
        headers=headers or DEFAULT_HEADERS,
        timeout=timeout,
        verify=False,
        allow_redirects=True,
    )
    resp.raise_for_status()
    if as_bytes:
        resp.raw  # noop; caller uses .content
    return resp


# ── xlsx 下载 + 解析 ──────────────────────────────────────────────────
def _download_xlsx(output_dir: str) -> str:
    """删除本地旧 xlsx 后重新下载，返回本地路径。"""
    local_path = os.path.join(output_dir, XLSX_FILENAME)
    if os.path.exists(local_path):
        try:
            os.remove(local_path)
            logger.info("[zdgksxml] removed stale local xlsx: %s", local_path)
        except OSError as e:
            logger.warning("[zdgksxml] failed to remove stale xlsx: %s", e)

    os.makedirs(output_dir, exist_ok=True)
    logger.info("[zdgksxml] downloading xlsx: %s", XLSX_URL)
    resp = _http_get(XLSX_URL, headers=DEFAULT_HEADERS, timeout=60)
    with open(local_path, "wb") as f:
        f.write(resp.content)
    logger.info("[zdgksxml] xlsx saved: %s (%d bytes)", local_path, len(resp.content))
    return local_path


def _build_merged_lookup(ws) -> Dict[int, int]:
    """构建 row → representative_row 映射（针对所有合并列）。

    若 row 落在某个 merged_range 内，映射到该 range 的 min_row。
    用于继承 col B/C/D（公开类别/二级/事项）的合并值。
    """
    lookup: Dict[int, int] = {}
    for mr in ws.merged_cells.ranges:
        for r in range(mr.min_row, mr.max_row + 1):
            # 多个 range 可能重叠（不应发生），保留最小 row
            prev = lookup.get(r, r)
            if mr.min_row < prev:
                lookup[r] = mr.min_row
    return lookup


def _parse_xlsx(xlsx_path: str) -> List[Dict[str, Any]]:
    """解析 xlsx，按 (matter, legal_doc_url) 分组聚合条款，返回 items 列表。"""
    import openpyxl  # 延迟导入，减少模块加载开销

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # 表头定位：R2 是列名，R1 是大标题
    header_row = 2
    col_by_key: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        raw = ws.cell(header_row, c).value
        if not raw:
            continue
        label = str(raw).strip().replace("\n", "")
        if label in _HEADER_MAP:
            col_by_key[_HEADER_MAP[label]] = c
    # 必须列
    required = ("seq_no", "category_l1", "matter", "legal_doc_title")
    missing = [k for k in required if k not in col_by_key]
    if missing:
        raise RuntimeError(f"[zdgksxml] xlsx missing required columns: {missing}")

    merged_lookup = _build_merged_lookup(ws)

    def cell_value(r: int, key: str) -> str:
        col = col_by_key.get(key)
        if not col:
            return ""
        rep = merged_lookup.get(r, r)
        v = ws.cell(rep, col).value
        return "" if v is None else str(v).strip()

    def hyperlink_url(r: int) -> str:
        col = col_by_key.get("legal_doc_title")
        if not col:
            return ""
        cell = ws.cell(r, col)
        if cell.hyperlink and cell.hyperlink.target:
            return str(cell.hyperlink.target).strip()
        return ""

    def cell_text(r: int, key: str) -> str:
        """普通单元格文本（不查合并），用于条款/内容等不合并的列。"""
        col = col_by_key.get(key)
        if not col:
            return ""
        v = ws.cell(r, col).value
        return "" if v is None else str(v).strip()

    # 按 (matter, legal_doc_url) 聚合
    grouped: Dict[Tuple[str, str], Dict[str, Any]] = {}
    order: List[Tuple[str, str]] = []  # 保留 xlsx 顺序

    for r in range(header_row + 1, ws.max_row + 1):
        url = hyperlink_url(r)
        if not url:
            continue  # 没有超链接的行（如分组合并行末尾的填充行）跳过
        matter = cell_value(r, "matter")
        if not matter:
            # 行内未填事项且不在合并范围内（罕见），尝试 back-fill
            for rr in range(r - 1, header_row, -1):
                v = cell_value(rr, "matter")
                if v:
                    matter = v
                    break
        key = (matter, url)
        if key not in grouped:
            order.append(key)
            grouped[key] = {
                "seq_no": cell_value(r, "seq_no"),
                "category_l1": cell_value(r, "category_l1"),
                "category_l2": cell_value(r, "category_l2"),
                "matter": matter,
                "disclosure_content": cell_text(r, "disclosure_content"),
                "legal_doc_title": cell_text(r, "legal_doc_title"),
                "legal_doc_url": url,
                "legal_doc_clause": cell_text(r, "legal_doc_clause"),
                "disclosure_deadline": cell_text(r, "disclosure_deadline"),
                "disclosure_period": cell_text(r, "disclosure_period"),
                "disclosure_subject": cell_text(r, "disclosure_subject"),
                "disclosure_duty": cell_text(r, "disclosure_duty"),
                "disclosure_method": cell_text(r, "disclosure_method"),
                "disclosure_channel": cell_text(r, "disclosure_channel"),
                "_source_row": r,
            }
        else:
            # 同 (matter, url) 多行：累加条款（多条款以空行分隔）
            item = grouped[key]
            extra_clause = cell_text(r, "legal_doc_clause")
            if extra_clause:
                if item["legal_doc_clause"]:
                    item["legal_doc_clause"] += "\n\n" + extra_clause
                else:
                    item["legal_doc_clause"] = extra_clause
            # 其它字段若 group 首行空则用本行补
            for k in ("disclosure_content", "disclosure_deadline",
                      "disclosure_period", "disclosure_subject",
                      "disclosure_duty", "disclosure_method",
                      "disclosure_channel"):
                if not item.get(k):
                    item[k] = cell_text(r, k)

    return [grouped[k] for k in order]


# ── 详情抓取 ──────────────────────────────────────────────────────────
def _decode_html(resp: requests.Response) -> str:
    """智能解码 HTML（utf-8 → gbk → gb2312 → latin-1 兜底）。"""
    raw = resp.content
    for enc in ("utf-8", "gbk", "gb2312", "latin-1"):
        try:
            return raw.decode(enc)
        except (UnicodeDecodeError, LookupError):
            continue
    return resp.text  # 最后兜底


def _extract_detail_markdown(html: str) -> str:
    """从 HTML 中识别正文并转换为 Markdown。

    选择器优先级见 DETAIL_SELECTORS；全部 miss 时回退到最大 <div> 或全 <p>。
    """
    from bs4 import BeautifulSoup
    try:
        from markdownify import markdownify as md
    except ImportError:
        md = None

    soup = BeautifulSoup(html, "html.parser")
    # 先去除 script/style/nav/footer，避免干扰
    for tag in soup(["script", "style", "nav", "footer", "header", "aside"]):
        tag.decompose()

    body_node = None
    for sel in DETAIL_SELECTORS:
        nodes = soup.select(sel)
        if nodes:
            # 取文本最长的那个（同名 selector 偶尔出现多次）
            body_node = max(nodes, key=lambda n: len(n.get_text(strip=True)))
            text_len = len(body_node.get_text(strip=True))
            if text_len >= DETAIL_CONTENT_MIN_LEN:
                break
            body_node = None  # 文本太短，继续尝试下一个 selector

    if body_node is None:
        # 回退：最大 <div>
        all_divs = soup.find_all("div")
        if all_divs:
            body_node = max(all_divs, key=lambda n: len(n.get_text(strip=True)))
        if not body_node or len(body_node.get_text(strip=True)) < DETAIL_CONTENT_MIN_LEN:
            # 最终回退：所有 <p>
            ps = soup.find_all("p")
            if ps:
                parts = []
                for p in ps:
                    t = p.get_text(strip=True)
                    if t:
                        parts.append(t)
                joined = "\n\n".join(parts)
                return joined[:8000] if joined else ""

    if not body_node:
        return ""

    # 转 Markdown
    if md is not None:
        try:
            md_text = md(str(body_node), heading_style="ATX")
            # 压缩连续 3+ 换行为 2 个
            md_text = re.sub(r"\n{3,}", "\n\n", md_text).strip()
            # 截断超长正文（保护 KB 上传）
            return md_text[:8000]
        except Exception as e:
            logger.warning("[zdgksxml] markdownify failed: %s, fallback to plain text", e)

    # 无 markdownify 时，纯文本兜底
    return body_node.get_text("\n", strip=True)[:8000]


def _fetch_detail(url: str) -> str:
    """抓取 URL 并返回 markdown 正文；失败返回空串。"""
    try:
        _throttle()
        resp = _http_get(url, headers=DETAIL_HEADERS, timeout=DETAIL_TIMEOUT)
        html = _decode_html(resp)
        if not html:
            return ""
        return _extract_detail_markdown(html)
    except Exception as e:
        logger.warning("[zdgksxml] detail fetch failed (%s): %s", url, e)
        return ""


# ── item 构造 ─────────────────────────────────────────────────────────
def _build_title(parsed: Dict[str, Any]) -> str:
    """标题 = 公开类别 / 公开事项 - 依据文件名。"""
    parts = []
    l1 = parsed.get("category_l1", "")
    matter = parsed.get("matter", "")
    if l1 and matter and l1 != matter:
        parts.append(f"{l1}/{matter}")
    elif matter:
        parts.append(matter)
    elif l1:
        parts.append(l1)
    title = parts[0] if parts else "未命名"
    doc_title = parsed.get("legal_doc_title", "").strip()
    # 去掉原文中常见的首尾换行/空白
    doc_title = re.sub(r"^\s+|\s+$", "", doc_title).replace("\n", " ")
    if doc_title:
        title = f"{title} - {doc_title}"
    return title[:500]  #crawler_result.title 长度限制


def _build_markdown(parsed: Dict[str, Any], detail_md: str) -> str:
    """把 13 列字段 + 详情正文拼成 markdown。"""
    lines: List[str] = []
    matter = parsed.get("matter", "")
    l1 = parsed.get("category_l1", "")
    l2 = parsed.get("category_l2", "")

    # 标题块
    header_path = " / ".join([p for p in (l1, l2, matter) if p])
    lines.append(f"# {header_path or '未命名事项'}")
    lines.append("")

    # 公开内容
    content = parsed.get("disclosure_content", "").strip()
    if content:
        lines.append("## 公开内容")
        lines.append("")
        lines.append(content)
        lines.append("")

    # 依据文件
    doc_title = parsed.get("legal_doc_title", "").strip()
    doc_url = parsed.get("legal_doc_url", "").strip()
    if doc_title or doc_url:
        lines.append("## 公开依据")
        lines.append("")
        if doc_title:
            lines.append(f"**文件：** {doc_title}")
        if doc_url:
            lines.append(f"**链接：** {doc_url}")
        lines.append("")

    # 条款
    clause = parsed.get("legal_doc_clause", "").strip()
    if clause:
        lines.append("## 公开依据文件条款")
        lines.append("")
        lines.append(clause)
        lines.append("")

    # 公开属性表
    prop_rows = [
        ("公开时限", parsed.get("disclosure_deadline", "")),
        ("公开期限", parsed.get("disclosure_period", "")),
        ("公开主体", parsed.get("disclosure_subject", "")),
        ("公开责任", parsed.get("disclosure_duty", "")),
        ("公开方式", parsed.get("disclosure_method", "")),
        ("公开渠道", parsed.get("disclosure_channel", "")),
    ]
    if any(v.strip() for _, v in prop_rows):
        lines.append("## 公开属性")
        lines.append("")
        lines.append("| 属性 | 内容 |")
        lines.append("| --- | --- |")
        for label, val in prop_rows:
            v = (val or "").strip().replace("\n", " ")
            if v:
                lines.append(f"| {label} | {v} |")
        lines.append("")

    # 详情页正文（外域法规原文）
    if detail_md:
        lines.append("---")
        lines.append("")
        lines.append("## 详情页正文（自动抓取）")
        lines.append("")
        lines.append(detail_md)
        lines.append("")

    return "\n".join(lines)


def _build_item_dict(parsed: Dict[str, Any], idx: int) -> Optional[Dict[str, Any]]:
    """把 xlsx 解析出的 parsed dict 转成 item_from_dict 可消费的 dict。

    idx 仅作为锚点 token 的最后兜底（seq_no 和 matter 都缺失时使用），
    不应被调用方视为稳定的 PK 组成部分。
    """
    legal_doc_url = parsed.get("legal_doc_url", "").strip()
    if not legal_doc_url:
        return None  # 没有超链接的行不能进库（PK 依赖 url）

    # 抓详情页正文
    detail_md = _fetch_detail(legal_doc_url)

    # source_url 加锚点让 PK 唯一（同一个法规文件被多个事项引用时各为独立条目）
    # 锚点必须用稳定值，不能用本次运行的顺序号 idx —— 否则 xlsx 行序变化
    # （新增/删除/重排）会让 idx 漂移，导致 result_id 变化、旧条目沦为孤儿、
    # 同一条目被反复当作新增写入。
    # 稳定值：seq_no（xlsx 序号列，政府目录的稳定标识）+ matter 短 hash 兜底。
    seq_no_raw = str(parsed.get("seq_no", "")).strip()
    matter_raw = (parsed.get("matter", "") or "").strip()
    if seq_no_raw:
        anchor_token = f"seq{seq_no_raw}"
    elif matter_raw:
        # 无序号时用 matter 的短 hash（前 8 位）保证稳定且唯一
        anchor_token = "m" + hashlib.md5(matter_raw.encode("utf-8")).hexdigest()[:8]
    else:
        # 最后兜底才用 idx（极端情况：序号和事项都缺失）
        anchor_token = f"i{idx}"
    source_url = f"{legal_doc_url}#matter-{anchor_token}"

    title = _build_title(parsed)
    content_md = _build_markdown(parsed, detail_md)

    return {
        # NormalizedItem 标准字段
        "title": title,
        "url": source_url,            # 含锚点，用于 PK
        "content": content_md,
        "date": "",                   # xlsx 无日期字段
        "news_type": NEWS_TYPE,
        "source_site": SITE_ID,
        "source": SOURCE_NAME,
        "section": parsed.get("category_l1", ""),
        # 扩展表 13 列字段（CollectionWriter._write_zdgksxml_ext 消费）
        "seq_no": parsed.get("seq_no", ""),
        "category_l1": parsed.get("category_l1", ""),
        "category_l2": parsed.get("category_l2", ""),
        "matter": parsed.get("matter", ""),
        "disclosure_content": parsed.get("disclosure_content", ""),
        "legal_doc_title": parsed.get("legal_doc_title", ""),
        "legal_doc_url": legal_doc_url,  # 真实 URL，不带锚点
        "legal_doc_clause": parsed.get("legal_doc_clause", ""),
        "disclosure_deadline": parsed.get("disclosure_deadline", ""),
        "disclosure_period": parsed.get("disclosure_period", ""),
        "disclosure_subject": parsed.get("disclosure_subject", ""),
        "disclosure_duty": parsed.get("disclosure_duty", ""),
        "disclosure_method": parsed.get("disclosure_method", ""),
        "disclosure_channel": parsed.get("disclosure_channel", ""),
    }


# ── 入口 ──────────────────────────────────────────────────────────────
def run(
    tenant_id: str,
    kb_id: str,
    task_name: str = "",
    task_id: str = "",
    writer_mode: str = "collection",
    category: str = "zdgksxml",
    date_filter: str = "",
    full_crawl: bool = False,
    force_run: bool = False,
    site_config: Any = None,
    output_dir: Optional[str] = None,
) -> Dict[str, Any]:
    """unified_crawler custom_runner 调度入口。

    Returns:
        与 CrawlerEngine.run 兼容的 summary dict:
        {"status", "pages", "items_found", "items_new", "items_updated",
         "kb_uploaded", "attachments_uploaded", "errors"}
    """
    try:
        # 注意: 必须用 common.settings (而非 rag.settings)。
        # common.settings 持有 STORAGE_IMPL 模块级变量，KBUploader 通过
        # FileService.upload_document 间接读取。rag.settings 没有 init_settings
        # 属性，且两者不是同一模块实例，调错会导致 STORAGE_IMPL=None，KB 上传
        # 全部失败 ('NoneType' object has no attribute 'obj_exist')。
        from common import settings as _settings
        _settings.init_settings()
    except Exception:
        pass

    import warnings
    warnings.filterwarnings("ignore", message="Workbook contains no default style")

    stats: Dict[str, Any] = {
        "status": "running",
        "pages": 1,
        "items_found": 0,
        "items_new": 0,
        "items_updated": 0,
        "kb_uploaded": 0,
        "attachments_uploaded": 0,
        "errors": [],
    }
    errors = stats["errors"]

    if not output_dir:
        output_dir = os.path.join(_PROJECT_ROOT, "rag", "zdgksxml_output")
    os.makedirs(output_dir, exist_ok=True)

    # 1. 下载 xlsx
    try:
        xlsx_path = _download_xlsx(output_dir)
    except Exception as e:
        logger.error("[zdgksxml] xlsx download failed: %s", e)
        stats["status"] = "error"
        errors.append(f"xlsx download failed: {e}")
        return stats

    # 2. 解析 xlsx → grouped items
    try:
        parsed_items = _parse_xlsx(xlsx_path)
    except Exception as e:
        logger.error("[zdgksxml] xlsx parse failed: %s", e)
        stats["status"] = "error"
        errors.append(f"xlsx parse failed: {e}")
        return stats

    stats["items_found"] = len(parsed_items)
    logger.info("[zdgksxml] parsed %d items from xlsx", len(parsed_items))
    if not parsed_items:
        stats["status"] = "success"
        return stats

    # 3. 初始化存储管道
    from rag.svr.crawler_engine.storage_pipeline import StoragePipeline
    pipeline = StoragePipeline(
        kb_id=kb_id,
        tenant_id=tenant_id,
        parser_id="naive",
        site_id=SITE_ID,
        task_name=task_name or SITE_ID,
        output_dir=output_dir,
        writer_mode="collection",
        category="zdgksxml",
        task_id=task_id,
        date_filter=date_filter or "",
        site_display=f"{SITE_NAME} {SITE_DOMAIN}",
    )

    # 4. 逐条处理
    processed = 0
    for idx, parsed in enumerate(parsed_items, 1):
        try:
            item_dict = _build_item_dict(parsed, idx)
            if not item_dict:
                continue
            normalized = item_from_dict(item_dict, site_id=SITE_ID,
                                        section=item_dict.get("section", ""))
            store_result = pipeline.store(normalized)
            if store_result.get("project_id"):
                processed += 1
            if idx % 20 == 0:
                logger.info("[zdgksxml] progress %d/%d (processed=%d)",
                            idx, len(parsed_items), processed)
        except Exception as e:
            logger.exception("[zdgksxml] item %d failed: %s", idx, e)
            errors.append(f"item#{idx} ({parsed.get('legal_doc_url','')[:80]}): {e}")

    # 5. 汇总 stats
    try:
        pipeline.cleanup()
    except Exception:
        pass

    cw_stats = pipeline._collection_writer.stats if pipeline._collection_writer else {}
    pl_stats = pipeline.stats
    stats["items_new"] = cw_stats.get("results_new", 0)
    stats["items_updated"] = cw_stats.get("results_updated", 0)
    stats["kb_uploaded"] = pl_stats.get("kb_uploaded", 0)
    stats["attachments_uploaded"] = pl_stats.get("attachments_uploaded", 0)
    stats["status"] = "success" if not errors else "success_with_errors"

    logger.info(
        "[zdgksxml] DONE: processed=%d, new=%d, updated=%d, kb=%d, errors=%d",
        processed, stats["items_new"], stats["items_updated"],
        stats["kb_uploaded"], len(errors),
    )
    # 触发脚本 (_trigger_*.py) 通过正则解析此行回写 crawler_task.last_run_*
    # 必须在 run() 末尾输出，否则 last_run_summary 会全为 0
    print(
        f"[CRAWLER] Done: {stats['items_new']} new items, "
        f"{stats['items_updated']} updated, {stats['kb_uploaded']} kb uploaded",
        flush=True,
    )
    return stats


# ── CLI 直跑（debug 用） ─────────────────────────────────────────────
if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    )
    import argparse
    p = argparse.ArgumentParser()
    p.add_argument("--tenant-id", required=True)
    p.add_argument("--kb-id", default="")
    p.add_argument("--task-name", default="manual-zdgksxml")
    p.add_argument("--task-id", default="")
    p.add_argument("--output-dir", default=None)
    args = p.parse_args()

    summary = run(
        tenant_id=args.tenant_id,
        kb_id=args.kb_id,
        task_name=args.task_name,
        task_id=args.task_id,
        output_dir=args.output_dir,
    )
    print("\n[CRAWLER] Done: "
          f"{summary['items_new']} new items, "
          f"{summary['items_updated']} updated, "
          f"{summary['kb_uploaded']} kb uploaded")
    sys.exit(0 if summary["status"] in ("success", "success_with_errors") else 1)
